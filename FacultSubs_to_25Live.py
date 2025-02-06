#FacultSubs_to_25Live
import pyautogui
import time
import csv
import webbrowser
import pandas as pd
import openpyxl
import os
import glob
from datetime import datetime, timedelta, time as datetime_time
import re
import time  # Import the time module for sleep

event_names_list = []
event_dates_list = []
start_times_list = []
durations_list = []
num_of_students_list = []
event_locations_list = []
end_times_list = []

# Function to allow for adding of test duration to start time for an end time
def calculate_end_time(start_time_str, duration_minutes):
    # Ensure start_time_str is a string
    if isinstance(start_time_str, datetime_time):  # This checks if it's a datetime.time object
        start_time_str = start_time_str.strftime('%I:%M %p')
    
    # Parse the start time string into a datetime object
    start_time = datetime.strptime(start_time_str, '%I:%M %p')
    
    # Add 15 minutes to the duration
    total_duration = duration_minutes + 15
    
    # Calculate the end time by adding the total duration to the start time
    end_time = start_time + timedelta(minutes=total_duration)
    
    # Format the end time back into a string in military time
    end_time_str = end_time.strftime('%H:%M')
    
    return end_time_str
    
def get_worksheet_title(file_path):
    file_extension = os.path.splitext(file_path)[1].lower()

    semester_season = input("Which semester? Enter SP for Spring or FL for Fall: ").upper()
    semester_year = input("Which year? Enter the last 2 digits: ")

    search_string = f"{semester_season}{semester_year}"

    if file_extension == '.csv':
        try:
            df = pd.read_csv(file_path, encoding='latin-1')
        except UnicodeDecodeError:
            print("Error: Unable to decode the file with 'latin-1' encoding.")
            return None

        df.columns = df.columns.str.upper()

        if any(search_string in col for col in df.columns):
            print(f"Found data for: {search_string}")
            return search_string
        else:
            print("Error: No data found with both the semester season and semester year in the columns.")
            return None

    elif file_extension in ['.xlsx', '.xlsm', '.xltx', '.xltm']:
        workbook = openpyxl.load_workbook(file_path)

        for sheet in workbook.sheetnames:
            if search_string in sheet:
                print(f"Found worksheet: {sheet}")
                return sheet

        print("Error: No worksheet found with both the semester season and semester year in the name.")
        return None

    else:
        print("Error: Unsupported file format. Please provide a .csv or .xlsx file.")
        return None

def determine_drop_columns(headers):
    if len(headers) == 13:
        return [5, 6, 8, 9, 10, 11, 12], 13
    elif len(headers) == 15:
        return [1, 7, 8, 10, 11, 12, 13, 14], 15
    elif len(headers) == 17:
        return [1, 3, 4, 5, 6, 10, 11, 12, 13, 14, 15, 16], 17
    else:
        return []

def filter_row(row, drop_columns):
    return [value for i, value in enumerate(row) if i not in drop_columns]

def load_subject_with_dropped_columns(file_path):
    subject = []
    tests = []
    r25_found = False
    outlook_found = False
    
    worksheet_title = get_worksheet_title(file_path)
    if worksheet_title is None:
        return None, None, None

    file_extension = os.path.splitext(file_path)[1].lower()

    try:
        if file_extension == '.csv':
            with open(file_path, mode='r', encoding='latin-1') as file:
                reader = csv.reader(file)
                next(reader)  # Skip the first row
                headers = next(reader)
                drop_columns, hlength = determine_drop_columns(headers)
                filtered_headers = filter_row(headers, drop_columns)

                for row in reader:
                    if len(row) > 2 and ("aptop" in str(row[0]) or "aptop" in str(row[2])):
                        #laptop_found = True  
                        continue
                    if len(row) > hlength-1 and row[hlength-1]: #Check R25
                        r25_found = True
                        #continue
                    if len(row) > hlength-2 and row[hlength-2]: #Check Outlook
                        outlook_found = True
                        #continue
                    filtered_row = filter_row(row, drop_columns)
                    filtered_row.extend([r25_found, outlook_found])  # Add flags to the list
                    # Map headers to filtered row values, including the new flags
                    full_headers = filtered_headers + ["R25_Found", "Outlook_Found"]  # Update headers
                    subject.append({full_headers[i]: filtered_row[i] for i in range(len(full_headers))})
                    tests.append(filtered_row)
                    
                    r25_found = False
                    outlook_found = False

        elif file_extension in ['.xlsx', '.xlsm', '.xltx', '.xltm']:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook[worksheet_title]
            headers = [cell.value for cell in sheet[2]]
            drop_columns, hlength = determine_drop_columns(headers)
            filtered_headers = filter_row(headers, drop_columns)

            for row in sheet.iter_rows(min_row=3, values_only=True):
                if len(row) > 2 and ("aptop" in str(row[0]) or "aptop" in str(row[2])):
                    #laptop_found = True  
                    continue
                if len(row) > hlength-2 and row[hlength-2]: #Check Outlook
                    outlook_found = True
                    #continue
                if len(row) > hlength-1 and row[hlength-1]: #Check R25
                    r25_found = True
                    #continue
                filtered_row = filter_row(row, drop_columns)
                filtered_row.extend([outlook_found, r25_found])  # Add flags to the list
                # Map headers to filtered row values, including the new flags
                full_headers = filtered_headers + ["Outlook_Found", "R25_Found"]  # Update headers
                subject.append({full_headers[i]: filtered_row[i] for i in range(len(full_headers))})
                tests.append(filtered_row)
                  
                r25_found = False
                outlook_found = False

        else:
            print("Error: Unsupported file format. Please provide a .csv or .xlsx file.")
            return None, None, None

    except Exception as e:
        print(f"Error processing file: {e}")
        return None, None, None

    return subject, tests, headers, worksheet_title

def parse_duration(duration_str):
    if duration_str:
        # Ensure duration_str is a string
        duration_str = str(duration_str)
        # Extract numeric part from the duration string
        return int(''.join(filter(str.isdigit, duration_str)))
    return None
    
def validate_tests(test, headers, title):
    if isinstance(test, dict):  # Check if test is a dictionary
        test = list(test.values())  # Convert dictionary values to a list
    outlook = test[-2]
    r25 = test[-1]
    temp = title.split()
    if "rt" in title.lower() or "dms" in title.lower():
        title = "[RT/DMS] "
    elif "day" in title.lower():
        title = "[N" + temp[0] + "] Day "
    elif "eve" in title.lower():
        title = "[N" + temp[0] + "] Eve "
    else:
        title = "[N" + temp[0] + "] "
        
    if len(headers) == 15 or len(headers) == 17:
        event_name = title + re.sub(r'[^A-Za-z0-9\s/&]', '',  str(test[0])) + ': ' + re.sub(r'[^A-Za-z0-9\s/&]', '',  str(test[1])) if len(test) > 1 and test[0] and test[1] else None
        event_date = test[2] if len(test) > 2 and test[2] else None
        start_time = test[3] if len(test) > 3 and test[3] else None
        duration = parse_duration(test[4]) if len(test) > 4 and test[4] else None
        num_of_students = test[5] if len(test) > 5 and test[5] else None
        event_location = test[6] if len(test) > 6 and test[6] else None
    elif len(headers) == 13:
        event_name = title + re.sub(r'[^A-Za-z0-9\s/&]', '',  str(test[0]))  if len(test) > 0 and test[0] else None
        event_date = test[1] if len(test) > 1 and test[1] else None
        start_time = test[2] if len(test) > 2 and test[2] else None
        duration = parse_duration(test[3]) if len(test) > 3 and test[3] else None
        num_of_students = test[4] if len(test) > 4 and test[4] else None
        event_location = test[5] if len(test) > 5 and test[5] else None
    else:
        return None  # Handle unexpected header lengths
    end_time = calculate_end_time(start_time, duration) if start_time and duration else None
    
    if event_name and event_date and start_time and duration and num_of_students and event_location and end_time:
        event_names_list.append(event_name)
        event_dates_list.append(event_date)
        start_times_list.append(start_time)
        durations_list.append(duration)
        num_of_students_list.append(num_of_students)
        event_locations_list.append(event_location)
        end_times_list.append(end_time)
    
    if not (event_name and event_date and start_time and duration and num_of_students and event_location):
        return None  # Indicate that this test should be removed
    
    return {
        'event_name': event_name,
        'event_date': event_date,
        'start_time': start_time,
        'duration': duration,
        'num_of_students': num_of_students,
        'event_location': event_location,
        'end_time': end_time,
        'Outlook_Input': outlook,
        'R25_Input': r25,
    }

def format_date_outlook(date_str):
    if not isinstance(date_str, datetime):
        try:
            # First attempt: Parse as 'd-%m-%Y' (e.g., 5-12-2025)
            date_obj = datetime.strptime(date_str, '%d-%m-%Y')
        except ValueError:
            try:
                # Second attempt: Parse as '%A, %B %d, %Y' (e.g., Monday, January 01, 2025)
                date_obj = datetime.strptime(date_str, '%A, %B %d, %Y')
            except ValueError:
                raise ValueError(f"event_date must be a datetime object or a valid date string. Got: {date_str}")
    else:
        date_obj = date_str
    
    return date_obj.strftime('%m/%d/%Y')
    
def format_date_r25(date_str):
    # Parse the input date string to a datetime object
    date_obj = datetime.strptime(date_str, '%A, %B %d, %Y')
    
    return date_obj.strftime('%m/%d/%Y')
    
# Function to open the R25 webpage
def open_R25():
    webbrowser.open("https://25live.collegenet.com/pro/jjc#!/home/event/form")
    print("Opened the 25Live CollegeNet website in the default browser.")
    # Delay for the website to load
    time.sleep(6)

def press_tabs(count):
    # Presses the Tab key `count` times with a delay.
    for _ in range(count):
        pyautogui.press('tab')
        time.sleep(0.2)
    print(f"Pressed Tab {count} times.")

# Function to open the R25 webpage
def open_outlook():
    webbrowser.open("https://outlook.office.com/calendar/view/month")
    print("Opened the Outlook website in the default browser.")
    # Delay for the website to load
    time.sleep(3)
    press_tabs(17) 

def navigate_and_enter_r25(key_presses):
    # Navigate and enter data based on key_presses
    for key, count, text in key_presses:
        # Perform the key navigation
        for _ in range(count):
            pyautogui.press(key)
            time.sleep(0.15)
        print(f"Pressed '{key}' {count} times.")

        # Enter text if provided
        if callable(text):  # Check if text is a function
            text()  # Execute the function
        elif isinstance(text, str):
            if text == "wait":
                time.sleep(3)
                print(f"Entered text: {text}ed for 3 seconds")
            else:
                pyautogui.write(text, interval=0.2)
                print(f"Entered text: {text}")
                
def insert_into_R25(valid_tests):
    if valid_tests.get('R25_Input'):
        print(f"Skipping {valid_tests.get('event_name')} as it is already in R25")
        return
    
    # Validate event_date
    if not valid_tests.get('event_date') or not isinstance(valid_tests.get('event_date'), (str, datetime)):
        print(f"Invalid or missing event_date: {valid_tests.get('event_date')}")
        return
    
    # Process and format event_date
    event_date = valid_tests.get('event_date')
    temp = format_date_outlook(valid_tests.get('event_date'))
    event_date = temp
    
    #Compare event_date with the specific date
    if (datetime.strptime(event_date, '%m/%d/%Y')) < datetime(2025, 1, 28):
        print(f"Skipping event on {event_date} as it is before {datetime(2025, 1, 28)}") #Checks if the event date has is in the past.
        return
    key_presses = [('', 0, "wait"),
    ('tab', 0, valid_tests['event_name']), ('tab', 2, ""), ('enter', 1, ""), ('down', 3, ""), ('enter', 1, ""),
    ('tab', 2, ""), ('enter', 1, ""), ('down', 1 if any(substring in valid_tests['event_name'] for substring in ["RADT", "DMS", "Para"]) else 2, ""), ('enter', 1, ""), ('tab', 2, str(valid_tests['num_of_students'])),
    ('tab', 5, str(valid_tests['event_name']) + ", " + str(valid_tests['duration']) + " Mins, Location: Room " + str(valid_tests['event_location']) + ", Students: " + str(valid_tests['num_of_students'])),
    ('tab', 3, event_date), ('tab', 1, ""), ('backspace', 9, str(valid_tests['start_time'])), ('tab', 1, ""), ('backspace', 9, str(valid_tests['end_time'])),
    ('tab', 51, ""), ('tab', 4, ""), ('enter', 1, ""), 
    ('', 0, "wait"),
    ('tab', 13, ""), ('enter' if "A" in valid_tests['event_location'] else 'tab', 3 if "A" not in valid_tests['event_location'] else 1, ""), ('tab', 3 if "A" in valid_tests['event_location'] else 0, ""), 
    ('', 0, "wait"),
    ('enter' if "B" in valid_tests['event_location'] else 'tab', 3 if "B" not in valid_tests['event_location'] else 1, ""), ('tab', 3 if "B" in valid_tests['event_location'] else 0, ""),                                                                  
    ('enter' if "C" in valid_tests['event_location'] else 'tab', 3 if "C" not in valid_tests['event_location'] else 1, ""), ('tab', 3 if "C" in valid_tests['event_location'] else 0, ""),  
    ('enter' if "D" in valid_tests['event_location'] else 'tab', 3 if "D" not in valid_tests['event_location'] else 1, ""), ('tab', 3 if "D" in valid_tests['event_location'] else 0, ""),
    ('tab', 11 if len(valid_tests['event_location']) == 1 else 20 if len(valid_tests['event_location']) == 2 else 29, ""), ('left', 1,  ""), ('tab', 1,  ""), ('left', 1,  ""),
    ('tab', 1,  ""), ('left', 1,  ""), ('tab', 14,  ""), ('enter', 1,  ""), ('down', 2,  ""),
    ('enter', 1,  ""), ('tab', 14,  ""), ('space', 1,  ""), ('tab', 5,  ""), 
    ('enter', 1, ""), ('', 0, "wait"),
    ('tab', 3,  ""), ('space', 1,  ""), ('tab', 1,  ""), ('enter', 1,  "")
]
    open_R25()
    navigate_and_enter_r25(key_presses)     
    
    time.sleep(7)

def navigate_and_enter_outlook(key_presses):
    # Navigate and enter data based on key_presses
    for key, count, text in key_presses:
        # Perform the key navigation
        for _ in range(count):
            pyautogui.press(key)
            time.sleep(0.15)
        print(f"Pressed '{key}' {count} times.")

        # Enter text if provided
        if callable(text):  # Check if text is a function
            text()  # Execute the function
        elif isinstance(text, str):
            if text == "wait":
                time.sleep(3)
                print(f"Entered text: {text}ed for 3 seconds")
            elif text == "Shift + tab":
                for _ in range(15):
                    pyautogui.keyDown('shift')  # Hold down the Shift key
                    pyautogui.press('tab')     # Press the Tab key
                    pyautogui.keyUp('shift')  # Release the Shift key
                pyautogui.press('enter')
            else:
                pyautogui.write(text, interval=0.1)
                print(f"Entered text: {text}")
                        
def insert_into_outlook(valid_tests):
    if valid_tests.get('Outlook_Input'):
        print(f"Skipping {valid_tests.get('event_name')} as it is already in Outlook")
        return
    
    # Validate event_date
    if not valid_tests.get('event_date') or not isinstance(valid_tests.get('event_date'), (str, datetime)):
        print(f"Invalid or missing event_date: {valid_tests.get('event_date')}")
        return
    
    # Process and format event_date
    temp = format_date_outlook(valid_tests.get('event_date'))
    event_date = temp
    
    #Compare event_date with the specific date
    if (datetime.strptime(event_date, '%m/%d/%Y')) < datetime(2025, 1, 28):
        print(f"Skipping event on {event_date} as it is before {datetime(2025, 1, 28)}")
        return
    
    # Validate and prepare key_presses
    event_name = valid_tests.get('event_name')
    end_time = valid_tests.get('end_time')
    start_time = (valid_tests.get('start_time')).strftime('%H:%M')
    
    if not event_name or not isinstance(event_name, str):
        print("Error: 'event_name' is missing or not a string.")
        return 
    if not end_time or not isinstance(end_time, str):
        print("Error: 'end_time' is missing or not a string.")
        return
        
    open_outlook() 
    key_presses = [
    ('', 0, "wait"), ('enter', 1,  event_name), ('tab', 2,  "Simeon.ngala"), ('enter', 1,  ""), ('tab', 2, event_date), ('tab', 1, start_time), ('tab', 1, end_time), ('tab', 6,  ""), ('space', 1,  ""), ('', 0,  "Shift + tab"), ('enter', 1,  "")
    ]
    
    # Execute key presses
    navigate_and_enter_outlook(key_presses)
    print(f"Test inserted into Outlook: {valid_tests}")
    time.sleep(5)

                
# Define the CSV/XLSX file and columns to select
for file in glob.glob("*.xlsx"):
    print(f"Processing file: {file}")
    # Call the function and load subject
    subject, tests, headers, title = load_subject_with_dropped_columns(file)

    if "day" in file.lower():
        title += " day "
    elif "eve" in file.lower():
        title += " eve"


    valid_tests = []
    valid_subject = []
    if subject:
        for sub in subject:
            result = validate_tests(sub, headers, title)  # Pass headers as an argument
            if result:
                valid_subject.append(result)
                insert_into_outlook(result)
                insert_into_R25(result)

    def old_way():            
        if not valid_subject:
            insert_into_outlook(valid_subject)
            print("Done inserting Class into Outlook")
            insert_into_R25(valid_subject)
            print("Done inserting Class into R25")
                    
